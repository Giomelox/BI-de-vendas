from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl import load_workbook
import pandas as pd

# Lendo a planilha BI frios
BI_semanal_frios = pd.read_excel('bi_semanal.xlsx', sheet_name = 0)

# Selecionando na coluna C (da folha de Frios), os clientes que possuem textos específicos
frios_formatado_Novo = BI_semanal_frios[BI_semanal_frios['Cliente'].str.contains(r'\b(?:NOVO ATACAREJO|NOVO ATACADO|CARPINA TREVO)\b', case = False, na = False)]

# Agrupando as colunas da planilha de acordo com a coluna C
frios_formatado_Novo = frios_formatado_Novo.sort_values('Cliente')

# Alterando o formato da data
frios_formatado_Novo['Data Vencimento'] = pd.to_datetime(frios_formatado_Novo['Data Vencimento'], format = '%d/%m/%Y').dt.strftime('%d/%m/%Y')

# Selecionando na coluna C (da folha de Frios), os clientes que possuem textos específicos
frios_formatado_MIX = BI_semanal_frios[BI_semanal_frios['Cliente'].str.contains(r'\b(?:MIX MATEUS|MATEUS)\b', case = False, na = False)]

# Agrupando as colunas da planilha de acordo com a coluna C
frios_formatado_MIX = frios_formatado_MIX.sort_values('Cliente')

# Alterando o formato da data
frios_formatado_MIX['Data Vencimento'] = pd.to_datetime(frios_formatado_MIX['Data Vencimento'], format = '%d/%m/%Y').dt.strftime('%d/%m/%Y')

# Lendo a planilha BI secos
BI_semanal_secos = pd.read_excel('bi_semanal.xlsx', sheet_name = 2)

# Selecionando a coluna C (da folha de Secos), os clientes que possuem textos específicos
secos_formatado_Novo = BI_semanal_secos[BI_semanal_secos['Cliente'].str.contains(r'\b(?:NOVO ATACAREJO|NOVO ATACADO|CARPINA TREVO)\b', case = False, na = False)]

# Agrupando as colunas da planilha de acordo com a coluna C
secos_formatado_Novo = secos_formatado_Novo.sort_values('Cliente')

# Alterando o formato da data
secos_formatado_Novo['Data Vencimento'] = pd.to_datetime(frios_formatado_MIX['Data Vencimento'], format = '%d/%m/%Y').dt.strftime('%d/%m/%Y')

# Selecionando na coluna C (da folha de Secos), os clientes que possuem textos específicos
secos_formatado_MIX = BI_semanal_secos[BI_semanal_secos['Cliente'].str.contains(r'\b(?:MIX MATEUS|MATEUS)\b', case = False, na = False)]

# Agrupando as colunas da planilha de acordo com a coluna C
secos_formatado_MIX = secos_formatado_MIX.sort_values('Cliente')

# Alterando o formato da data
secos_formatado_MIX['Data Vencimento'] = pd.to_datetime(frios_formatado_MIX['Data Vencimento'], format = '%d/%m/%Y').dt.strftime('%d/%m/%Y')

# Criando uma planilha só com os dataframes em folhas diferentes
with pd.ExcelWriter('BI Mix e Novo - Frios e Secos.xlsx') as writer:
    frios_formatado_Novo.to_excel(writer, sheet_name = 'Novo Frios', index = False)
    frios_formatado_MIX.to_excel(writer, sheet_name = 'Mix Frios', index = False)

    secos_formatado_Novo.to_excel(writer, sheet_name = 'Novo Secos', index = False)
    secos_formatado_MIX.to_excel(writer, sheet_name = 'Mix Secos', index = False)

def rebaixa_NOVO_frios():
    # Criar uma nova planilha para MIX frios, com rabaixas
    rebaixa_frios_NOVO = pd.DataFrame(columns = ['Stock Location Description', 'SKU Description', 'PLU', 'QTDE', 'PREÇO PDV', 'INVEST UND', 'PREÇO REBAIXA', 'VENC', 'SELL OUT', 'STATUS'])

    # Lendo a planilha principal formada e extraindo a folha de NOVO FRIOS
    df_BI_novo_frios = pd.read_excel('BI Mix e Novo - Frios e Secos.xlsx', sheet_name = 'Novo Frios')

    # Lendo a planilha de códigos para acrescentar na planilha de rebaixa
    df_codigo_produtos_frios_NOVO = pd.read_excel('cod frios novo.xlsx', sheet_name = 0)

    x = 2

    cliente_anterior = None

    # Adicionar as linhas do NOVO FRIOS na rebaixa
    for idx, row in df_BI_novo_frios.iterrows():
        cliente_atual = row['Cliente'].strip().lower()

        if cliente_anterior is None:
            cliente_anterior = cliente_atual

        # Adiciona a linha normalmente
        rebaixa_frios_NOVO.loc[len(rebaixa_frios_NOVO)] = {
            'Stock Location Description': row['Cliente'],
            'SKU Description': row['Produto Trade'],
            'PLU': '',
            'QTDE': row['Estoque (UN)'],
            'PREÇO PDV': '',
            'INVEST UND': f'=E{x} - G{x}',
            'PREÇO REBAIXA': '',
            'VENC': row['Data Vencimento'],
            'SELL OUT': f'=F{x} * D{x}',
            'STATUS': ''
        }

        x += 1

        # Se for o último ou o cliente mudar, adiciona bloco de separação
        proximo_cliente = (
            df_BI_novo_frios.iloc[idx + 1]['Cliente'].strip().lower()
            if idx + 1 < len(df_BI_novo_frios)
            else None
        )

        if cliente_atual != proximo_cliente:
            # Linha vazia
            rebaixa_frios_NOVO.loc[len(rebaixa_frios_NOVO)] = {col: '' for col in rebaixa_frios_NOVO.columns}
            
            x += 2

            # Linha com os títulos
            rebaixa_frios_NOVO.loc[len(rebaixa_frios_NOVO)] = {
                'Stock Location Description': 'Stock Location Description',
                'SKU Description': 'SKU Description',
                'PLU': 'PLU',
                'QTDE': 'QTDE',
                'PREÇO PDV': 'PREÇO PDV',
                'INVEST UND': 'INVEST UND',
                'PREÇO REBAIXA': 'PREÇO REBAIXA',
                'VENC': 'VENC',
                'SELL OUT': 'SELL OUT',
                'STATUS': 'STATUS'
            }

            cliente_anterior = None  # Reinicia para o próximo grupo
            
        # Extraindo da planilha de códigos, as respectivas informações dos produtos
        rebaixa_frios_NOVO['PLU'] = rebaixa_frios_NOVO['SKU Description'].map(df_codigo_produtos_frios_NOVO.set_index('SKU Description')['PLU'])
        rebaixa_frios_NOVO['PREÇO PDV'] = rebaixa_frios_NOVO['SKU Description'].map(df_codigo_produtos_frios_NOVO.set_index('SKU Description')['Custo'])

    # Salvando o dataframe como excel
    rebaixa_frios_NOVO.to_excel('REBAIXA FRIOS NOVO.xlsx', index = False)

    # Carregando o excel criado para alterar cores e dimensões
    workbook = load_workbook('REBAIXA FRIOS NOVO.xlsx')
    worksheet = workbook.active

    # Formatando as colunas de valores para REAL
    for row in worksheet.iter_rows(min_row = 2, max_row = worksheet.max_row):
        for cell in row:
            if cell.column_letter in ['E', 'F', 'G', 'I']:
                cell.number_format = 'R$ #,##0.00'

    # Alterando as cores das linhas
    for row in worksheet.iter_rows(min_row = 1, max_row = worksheet.max_row):
        if row[0].value == 'Stock Location Description':
            for cell in row:
                # Alterando a cor para cinza
                cell.fill = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')

                # Alterando o estilo da fonte para negrito
                cell.font = Font(bold = True)

    # Alterando a largura das linhas
    largura_colunas = [
        24.00, 9.71, 9.71, 48.14, 8.30, 10.29, 14.29, 11.14, 12.71, 10.60, 47.57
    ]

    for col_idx, largura in enumerate(largura_colunas, start = 1):
        col_letter = worksheet.cell(row = 1, column = col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = largura

    # Alterando a altura das linhas
    for x in range(len(rebaixa_frios_NOVO)):
        worksheet.row_dimensions[x].height = 15

    # Alinhando os itens nas linhas e adicionando bordas
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            cell.border = Border(
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin'),
            )

    # Salvando as alterações reescrevendo a planilha ja criada
    workbook.save('REBAIXA FRIOS NOVO.xlsx')

def rebaixa_NOVO_secos():
        # Criar uma nova planilha para MIX frios, com rabaixas
    rebaixa_secos_NOVO = pd.DataFrame(columns = ['Stock Location Description', 'SKU Description', 'PLU', 'QTDE', 'PREÇO PDV', 'INVEST UND', 'PREÇO REBAIXA',  'VENC', 'SELL OUT', 'STATUS'])

    # Lendo a planilha principal formada e extraindo a folha de NOVO FRIOS
    df_BI_novo_secos = pd.read_excel('BI Mix e Novo - Frios e Secos.xlsx', sheet_name = 'Novo Secos')

    # Lendo a planilha de códigos para acrescentar na planilha de rebaixa
    df_codigo_produtos_secos_NOVO = pd.read_excel('cod secos novo.xlsx', sheet_name = 0)

    x = 2

    cliente_anterior = None

    # Adicionar as linhas do NOVO FRIOS na rebaixa
    for idx, row in df_BI_novo_secos.iterrows():
        cliente_atual = row['Cliente'].strip().lower()

        if cliente_anterior is None:
            cliente_anterior = cliente_atual

        # Adiciona a linha normalmente
        rebaixa_secos_NOVO.loc[len(rebaixa_secos_NOVO)] = {
            'Stock Location Description': row['Cliente'],
            'SKU Description': row['Produto Trade'],
            'PLU': '',
            'QTDE': row['Estoque (UN)'],
            'PREÇO PDV': '',
            'INVEST UND': f'=F{x} - G{x}',
            'PREÇO REBAIXA': '',
            'VENC': row['Data Vencimento'],
            'SELL OUT': f'=E{x} * H{x}',
            'STATUS': ''
        }

        x += 1
        

        # Se for o último ou o cliente mudar, adiciona bloco de separação
        proximo_cliente = (
            df_BI_novo_secos.iloc[idx + 1]['Cliente'].strip().lower()
            if idx + 1 < len(df_BI_novo_secos)
            else None
        )

        if cliente_atual != proximo_cliente:
            # Linha vazia
            rebaixa_secos_NOVO.loc[len(rebaixa_secos_NOVO)] = {col: '' for col in rebaixa_secos_NOVO.columns}
            x += 2

            # Linha com os títulos
            rebaixa_secos_NOVO.loc[len(rebaixa_secos_NOVO)] = {
                'Stock Location Description': 'Stock Location Description',
                'SKU Description': 'SKU Description',
                'PLU': 'PLU',
                'QTDE': 'QTDE',
                'PREÇO PDV': 'PREÇO PDV',
                'INVEST UND': 'INVEST UND',
                'PREÇO REBAIXA': 'PREÇO REBAIXA',
                'VENC': 'VENC',
                'SELL OUT': 'SELL OUT',
                'STATUS': 'STATUS'
            }
            cliente_anterior = None  # Reinicia para o próximo grupo
            
            # Extraindo da planilha de códigos, as respectivas informações dos produtos
            rebaixa_secos_NOVO['SKU'] = rebaixa_secos_NOVO['SKU Description'].map(df_codigo_produtos_secos_NOVO.set_index('SKU Description')['SKU'])
            rebaixa_secos_NOVO['COD'] = rebaixa_secos_NOVO['SKU Description'].map(df_codigo_produtos_secos_NOVO.set_index('SKU Description')['COD'])
            rebaixa_secos_NOVO['PREÇO PDV'] = rebaixa_secos_NOVO['SKU Description'].map(df_codigo_produtos_secos_NOVO.set_index('SKU Description')['Custo'])
        
    # Salvando o dataframe como excel
    rebaixa_secos_NOVO.to_excel('REBAIXA SECOS NOVO.xlsx', index = False)

    # Carregando o excel criado para alterar cores e dimensões
    workbook = load_workbook('REBAIXA SECOS NOVO.xlsx')
    worksheet = workbook.active

    # Formatando as colunas de valores para REAL
    for row in worksheet.iter_rows(min_row = 2, max_row = worksheet.max_row):
        for cell in row:
            if cell.column_letter in ['F', 'G', 'H', 'I']:
                cell.number_format = 'R$ #,##0.00'

    # Alterando as cores das linhas
    for row in worksheet.iter_rows(min_row = 1, max_row = worksheet.max_row):
        if row[0].value == 'Stock Location Description':
            for cell in row:
                # Alterando a cor para cinza
                cell.fill = PatternFill(start_color = 'CCCCCC', end_color = 'CCCCCC', fill_type = 'solid')

                # Alterando o estilo da fonte para negrito
                cell.font = Font(bold = True)

    # Alterando a largura das linhas
    largura_colunas = [
        24.00, 9.71, 9.71, 48.14, 8.30, 10.29, 14.29, 11.14, 12.71, 10.60, 47.57
    ]

    for col_idx, largura in enumerate(largura_colunas, start = 1):
        col_letter = worksheet.cell(row = 1, column = col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = largura

    # Alterando a altura das linhas
    for x in range(len(rebaixa_secos_NOVO)):
        worksheet.row_dimensions[x].height = 15

    # Alinhando os itens nas linhas e adicionando bordas
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            cell.border = Border(
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin'),
            )

    # Salvando as alterações reescrevendo a planilha ja criada
    workbook.save('REBAIXA SECOS NOVO.xlsx')

def rebaixa_MIX_frios():
    # Criar uma nova planilha para MIX frios, com rabaixas
    rebaixa_frios_MIX = pd.DataFrame(columns = ['Stock Location Description', 'SKU', 'COD', 'SKU Description', 'QTDE', 'PREÇO PDV', 'PREÇO REBAIXA', 'INVEST UND', 'INVEST TOTAL', 'VENC', 'STATUS'])

    # Lendo a planilha principal formada e extraindo a folha de NOVO FRIOS
    df_BI_MIX_frios = pd.read_excel('BI Mix e Novo - Frios e Secos.xlsx', sheet_name = 'Mix Frios')

    # Lendo a planilha de códigos para acrescentar na planilha de rebaixa
    df_codigo_produtos_frios_MIX = pd.read_excel('cod frios mix.xlsx', sheet_name = 0)

    x = 2

    cliente_anterior = None

    # Adicionar as linhas do NOVO FRIOS na rebaixa
    for idx, row in df_BI_MIX_frios.iterrows():
        cliente_atual = row['Cliente'].strip().lower()

        if cliente_anterior is None:
            cliente_anterior = cliente_atual

        # Adiciona a linha normalmente
        rebaixa_frios_MIX.loc[len(rebaixa_frios_MIX)] = {
            'Stock Location Description': row['Cliente'],
            'SKU': '',
            'COD': '',
            'SKU Description': row['Produto Trade'],
            'QTDE': row['Estoque (UN)'],
            'PREÇO PDV': '',
            'PREÇO REBAIXA': '',
            'INVEST UND': f'=F{x} - G{x}',
            'INVEST TOTAL': f'=E{x} * H{x}',
            'VENC': row['Data Vencimento'],
            'STATUS': ''
        }
        x += 1

        # Se for o último ou o cliente mudar, adiciona bloco de separação
        proximo_cliente = (
            df_BI_MIX_frios.iloc[idx + 1]['Cliente'].strip().lower()
            if idx + 1 < len(df_BI_MIX_frios)
            else None
        )

        if cliente_atual != proximo_cliente:
            # Linha vazia
            rebaixa_frios_MIX.loc[len(rebaixa_frios_MIX)] = {col: '' for col in rebaixa_frios_MIX.columns}
            x += 2

            # Linha com os títulos
            rebaixa_frios_MIX.loc[len(rebaixa_frios_MIX)] = {
                'Stock Location Description': 'Stock Location Description',
                'SKU': 'SKU',
                'COD': 'COD',
                'SKU Description': 'SKU Description',
                'QTDE': 'QTDE',
                'PREÇO PDV': 'PREÇO PDV',
                'PREÇO REBAIXA': 'PREÇO REBAIXA',
                'INVEST UND': 'INVEST UND',
                'INVEST TOTAL': 'INVEST TOTAL',
                'VENC': 'VENC',
                'STATUS': 'STATUS'
            }
            cliente_anterior = None  # Reinicia para o próximo grupo
            
            # Extraindo da planilha de códigos, as respectivas informações dos produtos
            rebaixa_frios_MIX['SKU'] = rebaixa_frios_MIX['SKU Description'].map(df_codigo_produtos_frios_MIX.set_index('SKU Description')['SKU'])
            rebaixa_frios_MIX['COD'] = rebaixa_frios_MIX['SKU Description'].map(df_codigo_produtos_frios_MIX.set_index('SKU Description')['COD'])
            rebaixa_frios_MIX['PREÇO PDV'] = rebaixa_frios_MIX['SKU Description'].map(df_codigo_produtos_frios_MIX.set_index('SKU Description')['Custo'])
        
    # Salvando o dataframe como excel
    rebaixa_frios_MIX.to_excel('REBAIXA FRIOS MIX.xlsx', index = False)

    # Carregando o excel criado para alterar cores e dimensões
    workbook = load_workbook('REBAIXA FRIOS MIX.xlsx')
    worksheet = workbook.active

    # Formatando as colunas de valores para REAL
    for row in worksheet.iter_rows(min_row = 2, max_row = worksheet.max_row):
        for cell in row:
            if cell.column_letter in ['F', 'G', 'H', 'I']:
                cell.number_format = 'R$ #,##0.00'

    # Alterando as cores das linhas
    for row in worksheet.iter_rows(min_row = 1, max_row = worksheet.max_row):
        if row[0].value == 'Stock Location Description':
            for cell in row:
                # Alterando a cor para cinza
                cell.fill = PatternFill(start_color = 'CCCCCC', end_color = 'CCCCCC', fill_type = 'solid')

                # Alterando o estilo da fonte para negrito
                cell.font = Font(bold = True)

    # Alterando a largura das linhas
    largura_colunas = [
        24.00, 9.71, 9.71, 48.14, 8.30, 10.29, 14.29, 11.14, 12.71, 10.60, 47.57
    ]

    for col_idx, largura in enumerate(largura_colunas, start = 1):
        col_letter = worksheet.cell(row = 1, column = col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = largura

    # Alterando a altura das linhas
    for x in range(len(rebaixa_frios_MIX)):
        worksheet.row_dimensions[x].height = 15

    # Alinhando os itens nas linhas e adicionando bordas
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            cell.border = Border(
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin'),
            )

    # Salvando as alterações reescrevendo a planilha ja criada
    workbook.save('REBAIXA FRIOS MIX.xlsx')

def rebaixa_MIX_secos():    
    
    # Criar uma nova planilha para MIX frios, com rabaixas
    rebaixa_secos_MIX = pd.DataFrame(columns = ['Stock Location Description', 'SKU', 'COD', 'SKU Description', 'QTDE', 'PREÇO PDV', 'PREÇO REBAIXA', 'INVEST UND', 'INVEST TOTAL', 'VENC', 'STATUS'])

    # Lendo a planilha principal formada e extraindo a folha de NOVO FRIOS
    df_BI_MIX_secos = pd.read_excel('BI Mix e Novo - Frios e Secos.xlsx', sheet_name = 'Mix Secos')

    # Lendo a planilha de códigos para acrescentar na planilha de rebaixa
    df_codigo_produtos_secos_MIX = pd.read_excel('cod secos mix.xlsx', sheet_name = 0)

    x = 2

    cliente_anterior = None

    # Adicionar as linhas do NOVO FRIOS na rebaixa
    for idx, row in df_BI_MIX_secos.iterrows():
        cliente_atual = row['Cliente'].strip().lower()

        if cliente_anterior is None:
            cliente_anterior = cliente_atual

        # Adiciona a linha normalmente
        rebaixa_secos_MIX.loc[len(rebaixa_secos_MIX)] = {
            'Stock Location Description': row['Cliente'],
            'SKU': '',
            'COD': '',
            'SKU Description': row['Produto Trade'],
            'QTDE': row['Estoque (UN)'],
            'PREÇO PDV': '',
            'PREÇO REBAIXA': '',
            'INVEST UND': f'=F{x} - G{x}',
            'INVEST TOTAL': f'=E{x} * H{x}',
            'VENC': row['Data Vencimento'],
            'STATUS': ''
        }
        x += 1

        # Se for o último ou o cliente mudar, adiciona bloco de separação
        proximo_cliente = (
            df_BI_MIX_secos.iloc[idx + 1]['Cliente'].strip().lower()
            if idx + 1 < len(df_BI_MIX_secos)
            else None
        )

        if cliente_atual != proximo_cliente:
            # Linha vazia
            rebaixa_secos_MIX.loc[len(rebaixa_secos_MIX)] = {col: '' for col in rebaixa_secos_MIX.columns}
            x += 2

            # Linha com os títulos
            rebaixa_secos_MIX.loc[len(rebaixa_secos_MIX)] = {
                'Stock Location Description': 'Stock Location Description',
                'SKU': 'SKU',
                'COD': 'COD',
                'SKU Description': 'SKU Description',
                'QTDE': 'QTDE',
                'PREÇO PDV': 'PREÇO PDV',
                'PREÇO REBAIXA': 'PREÇO REBAIXA',
                'INVEST UND': 'INVEST UND',
                'INVEST TOTAL': 'INVEST TOTAL',
                'VENC': 'VENC',
                'STATUS': 'STATUS'
            }
            cliente_anterior = None  # Reinicia para o próximo grupo
            
            # Extraindo da planilha de códigos, as respectivas informações dos produtos
            rebaixa_secos_MIX['SKU'] = rebaixa_secos_MIX['SKU Description'].map(df_codigo_produtos_secos_MIX.set_index('SKU Description')['SKU'])
            rebaixa_secos_MIX['COD'] = rebaixa_secos_MIX['SKU Description'].map(df_codigo_produtos_secos_MIX.set_index('SKU Description')['COD'])
            rebaixa_secos_MIX['PREÇO PDV'] = rebaixa_secos_MIX['SKU Description'].map(df_codigo_produtos_secos_MIX.set_index('SKU Description')['Custo'])
        
    # Salvando o dataframe como excel
    rebaixa_secos_MIX.to_excel('REBAIXA SECOS MIX.xlsx', index = False)

    # Carregando o excel criado para alterar cores e dimensões
    workbook = load_workbook('REBAIXA SECOS MIX.xlsx')
    worksheet = workbook.active

    # Formatando as colunas de valores para REAL
    for row in worksheet.iter_rows(min_row = 2, max_row = worksheet.max_row):
        for cell in row:
            if cell.column_letter in ['F', 'G', 'H', 'I']:
                cell.number_format = 'R$ #,##0.00'

    # Alterando as cores das linhas
    for row in worksheet.iter_rows(min_row = 1, max_row = worksheet.max_row):
        if row[0].value == 'Stock Location Description':
            for cell in row:
                # Alterando a cor para cinza
                cell.fill = PatternFill(start_color = 'CCCCCC', end_color = 'CCCCCC', fill_type = 'solid')

                # Alterando o estilo da fonte para negrito
                cell.font = Font(bold = True)

    # Alterando a largura das linhas
    largura_colunas = [
        24.00, 9.71, 9.71, 48.14, 8.30, 10.29, 14.29, 11.14, 12.71, 10.60, 47.57
    ]

    for col_idx, largura in enumerate(largura_colunas, start = 1):
        col_letter = worksheet.cell(row = 1, column = col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = largura

    # Alterando a altura das linhas
    for x in range(len(rebaixa_secos_MIX)):
        worksheet.row_dimensions[x].height = 15

    # Alinhando os itens nas linhas e adicionando bordas
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

            cell.border = Border(
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                top = Side(style = 'thin'),
                bottom = Side(style = 'thin'),
            )

    # Salvando as alterações reescrevendo a planilha ja criada
    workbook.save('REBAIXA SECOS MIX.xlsx')

rebaixa_NOVO_frios()
rebaixa_MIX_frios()
rebaixa_MIX_secos()